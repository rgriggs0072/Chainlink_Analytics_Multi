import streamlit as st
import pandas as pd
import snowflake.connector
import openpyxl
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
from PIL import Image
from db_utils.snowflake_utils import (
    get_snowflake_connection, get_snowflake_toml, upload_distro_grid_to_snowflake
)
from Formatting.DG_misc_format import DG_Misc_format
from utils.util import validate_chain_name

# Set the logo and header
def add_logo(logo_path, width, height):
    logo = Image.open(logo_path)
    return logo.resize((width, height))

my_logo = add_logo(logo_path="./images/Delta_Pacific/DeltaPacific_Logo.jpg", width=200, height=100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")
st.subheader("Misc Distribution Grid Processing")
st.write("[Download Distro Grid Template](https://github.com/rgriggs0072/ChainLinkAnalytics/raw/master/import_templates/Distribution_Grid_Template.xlsx)")
st.markdown("<hr>", unsafe_allow_html=True)


# Function to get chain options from Snowflake for dropdown
def get_options():
    toml_info = st.session_state.get('toml_info')
    if not toml_info:
        st.error("TOML information is not available. Please check the tenant ID and try again.")
        return []
    
    conn_toml = get_snowflake_toml(toml_info)
    cursor = conn_toml.cursor()
    cursor.execute('SELECT option_name FROM options_table ORDER BY option_name')
    
    return [row[0] for row in cursor]

    

# Retrieve options from Snowflake
options = get_options()

# Initialize session state variables
if 'new_option' not in st.session_state:
    st.session_state.new_option = ""
if 'option_added' not in st.session_state:
    st.session_state.option_added = False

# Create dropdown to select chain for formatting
if not options:
    st.warning("No options available. Please add options to the list.")
else:
    selected_option = st.selectbox('Select the Chain Distro Grid to format', options + ['Add new option...'], key="existing_option")
    st.session_state.selected_option = selected_option

if selected_option == 'Add new option...':
    new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
    if st.form_submit_button('Add Option') and new_option:
        options.append(new_option)
        toml_info = st.session_state.get('toml_info')
        conn_toml = get_snowflake_toml(toml_info)
        cursor = conn_toml.cursor()
        cursor.execute('INSERT INTO options_table (option_name) VALUES (%s)', (new_option,))
        conn_toml.commit()
        st.success('Option added successfully!')
        st.session_state.new_option = ""
else:
    st.write(f"You selected: {selected_option}")

# ======================================================================================================================
# Formatting Distribution Grid Spreadsheet Section
# ======================================================================================================================



# File upload for formatting
uploaded_file = st.file_uploader("Browse or drag here the Distribution Grid to Format", type=["xlsx"], key="uploaded_file")

# If a file is uploaded
if uploaded_file:
    # Validate the chain name right after file upload
    validate_chain_name(uploaded_file, st.session_state.selected_option)

    # If validation passes, update session state and proceed
    st.session_state['file_uploaded'] = True
    st.success("Chain name in the uploaded file matches the selected chain.")
    
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # If the "Reformat DG Spreadsheet" button is clicked
    if st.button("Reformat DG Spreadsheet"):
        formatted_workbook = None
        stream = BytesIO()

        # Format the non-pivot table
        formatted_workbook = DG_Misc_format(workbook, stream, st.session_state.selected_option)

        # Continue with further processing of the formatted workbook if needed
        st.success("Spreadsheet reformatted successfully!")
       


        # Allow file download after formatting is successful
        if formatted_workbook is not None:
            # Remove auto-filters if present
            for sheet_name in formatted_workbook.sheetnames:
                sheet = formatted_workbook[sheet_name]
                if sheet.auto_filter:
                    sheet.auto_filter.ref = None

            # Save the formatted workbook to a new stream for download
            stream = BytesIO()
            formatted_workbook.save(stream)
            stream.seek(0)

            st.download_button(
                label="Download formatted spreadsheet",
                data=stream.read(),
                file_name="formatted_spreadsheet.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.warning("No file has been prepared for download.")


# ======================================================================================================================
# Uploading Formatted Data to Snowflake Section
# ======================================================================================================================

# Section for uploading formatted data to Snowflake
st.markdown("<hr>", unsafe_allow_html=True)
st.subheader("Write Distribution Grid to Snowflake Utility")

# Create file uploader for uploading formatted spreadsheets
uploaded_files = st.file_uploader("Browse or select formatted Distribution Grid excel sheets", type=["xlsx"], accept_multiple_files=True)

for uploaded_file in uploaded_files:
    excel_file = pd.ExcelFile(uploaded_file)
    sheet_names = excel_file.sheet_names

    # Process each sheet in the uploaded Excel file
    for sheet_name in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        df.columns = df.columns.str.strip()

        # Validate chain_name in the uploaded file against the selected option
        if 'CHAIN_NAME' in df.columns:
            file_chain_name = df['CHAIN_NAME'].iloc[0].strip()
            if file_chain_name.lower() != selected_option.lower():
                st.error(f"Chain name in the file ({file_chain_name}) does not match the selected option ({selected_option}). Please correct the file.")
                continue
        else:
            st.error(f"'CHAIN_NAME' column not found in the uploaded file ({uploaded_file.name}). Please make sure the file has the correct format.")
            continue
       
        # Prepare to write to Snowflake
        button_key = f"import_button_{uploaded_file.name}_{sheet_name}"
        if st.button("Import Distro Grid into Snowflake", key=button_key):
            with st.spinner('Uploading data to Snowflake...'):
                upload_distro_grid_to_snowflake(df, selected_option, None)  # Assuming `update_spinner` is optional
                


# ======================================================================================================================
# Download Existing Distribution Grid Data from Snowflake Section
# ======================================================================================================================

if st.button("Prepare Distro Grid Spreadsheet for Selected Chain"):
    toml_info = st.session_state.get('toml_info')
    if not toml_info:
        st.error("TOML information is not available. Please check the tenant ID and try again.")
    else:
        conn_toml = get_snowflake_toml(toml_info)
        query = f"SELECT * FROM DISTRO_GRID WHERE CHAIN_NAME = '{selected_option}'"

        
        existing_data = pd.read_sql(query, conn_toml)
        

        if existing_data.empty:
            st.warning(f"No existing data found for chain '{selected_option}'.")
        else:
            download_stream = BytesIO()
            with pd.ExcelWriter(download_stream, engine='openpyxl') as writer:
                existing_data.to_excel(writer, sheet_name='Distro Grid Data', index=False)
            download_stream.seek(0)

            st.download_button(
                label=f"Download Distro Grid Data for {selected_option}",
                data=download_stream,
                file_name=f"{selected_option}_Distro_Grid.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
