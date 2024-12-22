import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle
from openpyxl import Workbook
from datetime import datetime, time
from dateutil.parser import parse
import re



#======================================================================================================================================
# This function will format the reset schedule to prepare for final integration into the reset_schedule table in Snowflake
#======================================================================================================================================
 
def format_RESET_schedule(workbook):
    # Get the worksheet by name ('Sheet1')
    ws = workbook['RESET_SCHEDULE_TEMPLATE']
    
    

    # # Fill column A with "CHAIN NAME SELECTED" starting from row 2
    # for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    #     for cell in row:
    #         cell.value = "FOOD MAXX"

    # Format column J as date "mm/dd/yyyy"
    date_format = NamedStyle(name='date_format', number_format='mm/dd/yyyy')
    for cell in ws['J']:
        cell.style = date_format
        

    # Loop through all cells in Column B
    for cell in ws['B']:
        # Check if the cell has a value (not empty)
        if cell.value:
            pass  # You can leave it empty or add specific logic for non-empty cells
        else:
            empty_cell_found = True
            st.warning("There is an empty Cell in column B (STORE_NUMBER). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
            return None  # Stop processing and return None when an empty cell is found

    # Initialize a flag to check for an empty cell
    empty_cell_found = False

    # Loop through all cells in Column C
    for cell in ws['C']:
        # Check if the cell has a value (not empty)
        if cell.value:
            # Convert the value to uppercase and remove leading/trailing spaces
            cell.value = cell.value.strip().upper()
        else:
            empty_cell_found = True
            st.warning("There is an empty Cell in column C (STORE_NAME). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
            return None  # Stop processing and return None when an empty cell is found
        
  




    # Initialize a flag to check for an empty cell
    empty_cell_found = False

    # Initialize a flag to check for an invalid date
    invalid_date_found = False

    # Regular expression pattern for date format (either 'mm/dd/yyyy' or 'm/d/yyyy')
    date_pattern = r'(\d{1,2}/\d{1,2}/\d{4})'

    # Loop through all cells in Column J, starting from row 2
    for cell in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        # Check if the cell has a value (not empty)
        if cell[0].value:
            # Check if the cell value is already a datetime object
            if not isinstance(cell[0].value, datetime):
                cell_value = str(cell[0].value)
                if not re.match(date_pattern, cell_value):
                    invalid_date_found = True
                    print(f"Invalid date format: {cell_value}")  # Print the invalid date value
                    break
        else:
            empty_cell_found = True
            print("Empty cell found")

    if invalid_date_found:
        st.warning("Invalid date format in column J (RESET_DATE). Please ensure all dates are correctly formatted as 'mm/dd/yyyy' or 'm/d/yyyy' and reformat the data.")
        return None  # Stop processing and return None when an invalid date is found

    if empty_cell_found:
        st.warning("There is an empty Cell in column J (RESET_DATE). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
        return None  # Stop processing and return None when an empty cell is found








    # Initialize a flag to check for an empty cell
    empty_cell_found = False

    # Loop through all cells in Column K (time)
    for cell in ws['K']:
        # Check if the cell has a value (not empty)
        if cell.value:
            pass
        else:
            empty_cell_found = True
            st.warning("There is an empty Cell in column K (TIME). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
            return None  # Stop processing and return None when an empty cell is found

    # Rename Columns as required to meet the objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')

    # Return the workbook if no empty cells are found
    return workbook


#======================================================================================================================================
#END of function will format the reset schedule to prepare for final integration into the reset_schedule table in Snowflake
#======================================================================================================================================






