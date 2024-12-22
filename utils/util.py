import streamlit as st
import bcrypt
from db_utils.snowflake_utils import get_snowflake_connection
import secrets
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from datetime import datetime, timedelta, timezone, time    
from dateutil.parser import parse
import pandas as pd
import numpy as np
from io import DEFAULT_BUFFER_SIZE, BytesIO
from PIL import Image
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from openpyxl.styles import Border, Side, PatternFill, Font, NamedStyle

# utils.util.py
import pandas as pd
import streamlit as st

import pandas as pd
import streamlit as st

def validate_chain_name(uploaded_file, selected_chain):
    """
    Validate that the chain name in the uploaded file matches the selected chain.
    
    Args:
        uploaded_file: The uploaded Excel file (Streamlit file uploader object).
        selected_chain: The chain selected by the user from the dropdown.
        
    """
    try:
        # Read the first sheet of the Excel file
        df = pd.read_excel(uploaded_file, sheet_name=0)

        # Ensure column names are stripped of whitespace
        df.columns = df.columns.str.strip()

        # Check if the CHAIN_NAME column exists
        if 'CHAIN_NAME' not in df.columns:
            st.error(f"'CHAIN_NAME' column not found in the uploaded file ({uploaded_file.name}). Please make sure the file has the correct format.")
            st.stop()  # Stop execution to prevent further processing

        # Get the chain name from the file and strip any leading/trailing spaces
        file_chain_name = df['CHAIN_NAME'].iloc[0].strip()

        # Compare chain name with the selected option
        if file_chain_name.lower() != selected_chain.lower():
            st.error(f"Chain name in the file ({file_chain_name}) does not match the selected option ({selected_chain}). Please correct the file or select the correct chain.")
            st.stop()  # Stop execution to prevent further processing

    except Exception as e:
        st.error(f"An error occurred while validating the chain name: {str(e)}")
        st.stop()  # Stop execution in case of any other errors




# Function to reset the password
def reset_password(username, reset_token, new_password, confirm_password):
    if new_password != confirm_password:
        st.error("The new passwords do not match. Please try again.")
        return False

    conn = get_snowflake_connection()
    cursor = conn.cursor()

    # Debugging log: check the received token
    st.write(f"Reset token received: {reset_token}")

    try:
        cursor.execute("""
            SELECT USERNAME, TOKEN_EXPIRY FROM USERDATA 
            WHERE RESET_TOKEN = %s AND TOKEN_EXPIRY > CURRENT_TIMESTAMP 
        """, (reset_token,))
        
        user_info = cursor.fetchone()

        if user_info:
            st.write(f"User info: {user_info}")  # Debugging log

            # Hash the new password
            hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

            # Update the user's password and clear the reset token and expiry
            cursor.execute("""
                UPDATE USERDATA SET HASHED_PASSWORD = %s, RESET_TOKEN = NULL, TOKEN_EXPIRY = NULL 
                WHERE RESET_TOKEN = %s
            """, (hashed_password, reset_token))
            conn.commit()

            return True  # Password reset was successful
        else:
            st.error("Invalid or expired reset token.")
            st.write(f"Invalid token or token expired")  # Debugging log
            return False  # Token was invalid or expired
    except Exception as e:
        st.write(f"Error during token validation: {e}")
    finally:
        cursor.close()
        conn.close()


# Function to generate a reset token
from datetime import datetime, timedelta, timezone

def generate_token(email):
    token = secrets.token_urlsafe()
    conn = get_snowflake_connection()
    cursor = conn.cursor()

    try:
        # Set the expiry to 1 hour from now in UTC
        expiry_time = datetime.now(timezone.utc) + timedelta(hours=1)

        # Store the token and its expiry time in Snowflake
        cursor.execute("""
            UPDATE USERDATA SET reset_token = %s, token_expiry = %s WHERE email = %s
        """, (token, expiry_time, email))
        
        conn.commit()
        st.write(f"Token saved: {token}, Expiry: {expiry_time}")  # Debugging log
    except Exception as e:
        st.write(f"Error saving token: {e}")
    finally:
        cursor.close()
        conn.close()

    return token


# Function to send reset link email
def send_reset_link(email, token):
    mailjet_creds = st.secrets["mailjet"]
    sender_email = "randy@chainlinkanalytics.com"
    smtp_username = mailjet_creds["API_KEY"]
    smtp_password = mailjet_creds["SECRET_KEY"]
    smtp_server = "in-v3.mailjet.com"
    smtp_port = 587

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = email
    msg['Subject'] = "Password Reset Request"
    html = f"""\ 
    <html>
      <body>
        <h3>You requested a password reset</h3>
        <p>Please use the following link to reset your password:</p>
        <p><a href="http://localhost:8501/reset_password?token={token}">Reset Password</a></p>
      </body>
    </html>
    """
    msg.attach(MIMEText(html, 'html'))

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_username, smtp_password)
            server.send_message(msg)
    except Exception as e:
        st.error(f"Failed to send reset email: {e}")




def create_user(username, email, fname, lname, selected_role, initial_status='Pending'):
    conn = get_snowflake_connection()
    cursor = conn.cursor()
    #print("I am create_user function now big daddy")
    try:
        # Assuming st.session_state['tenant_id'] is set during the login process
        if 'tenant_id' in st.session_state:
            admin_tenant_id = st.session_state['tenant_id']
        else:
            raise Exception("Tenant ID for the current session not found.")

        if not admin_tenant_id:
            raise Exception("Administrator tenant ID not found.")

        # Generate a temporary hashed password
        temp_password = "temporaryPassword"  # This should be randomly generated for each user
        hashed_temp_password = bcrypt.hashpw(temp_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

        # Insert the new user into the userdata table with the same tenant_id as the administrator
        cursor.execute("""
            INSERT INTO TENANTUSERDB.CHAINLINK_SCH.USERDATA (USER_ID, username, email, hashed_password, first_name, last_name, account_status, tenant_id)
            VALUES (TENANTUSERDB.CHAINLINK_SCH.USER_ID_SEQ.NEXTVAL, %s, %s, %s, %s, %s, %s, %s);
        """, (username, email, hashed_temp_password, fname, lname, initial_status, admin_tenant_id))

        # Fetch the USER_ID of the just inserted user
        cursor.execute("""
            SELECT USER_ID FROM USERDATA WHERE email = %s;
        """, (email,))
        user_id = cursor.fetchone()[0]  # Fetch the USER_ID

        # Get the ROLE_ID for the given role_name
        cursor.execute("""
            SELECT ROLE_ID FROM roles WHERE role_name = %s;
        """, (selected_role,))
        role_id = cursor.fetchone()[0]

        # Link the new user with the role in user_roles table
        cursor.execute("""
            INSERT INTO USER_ROLES (USER_ID, ROLE_ID)
            VALUES (%s, %s)
        """, (user_id, role_id))

        conn.commit()
        return True

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        conn.rollback()  # Rollback in case of error
        return False

    finally:
        cursor.close()
        conn.close()


def format_non_pivot_table(workbook, stream, selected_option):
    # Convert the worksheet to a DataFrame
    df = pd.DataFrame(workbook.active.values)

    # Initialize list to collect rows with missing values
    rows_with_missing_values = []

    # Iterate over each row to check for missing values in specific columns
    for index, row in df.iterrows():
        missing_columns = []
        for col_idx, value in enumerate(row):
            if pd.isna(value):
                column_name = None
                if col_idx == 0:
                    column_name = "STORE NAME"
                elif col_idx == 1:
                    column_name = "STORE NUMBER"
                elif col_idx == 2:
                    column_name = "UPC"

                if column_name:
                    missing_columns.append(column_name)

        if missing_columns:
            missing_columns_str = ", ".join(missing_columns)
            rows_with_missing_values.append(f"Row {index + 1}: {missing_columns_str}")


    # Display warnings in a modal-like section
    if rows_with_missing_values:
        with st.expander("Warning! Missing Values Detected", expanded=True):
            for warning in rows_with_missing_values:
                # Use st.error to make the warning messages red
                st.error(warning)
            
            if st.button("Acknowledge and Continue", key="acknowledge_warnings"):
                # This could reset a session state variable or perform some action to acknowledge the warnings
                st.session_state['acknowledged_warnings'] = True
                # Optionally rerun the app to refresh the state after acknowledgment
                st.rerun()

    # If everything is valid, you can continue processing the workbook
    return workbook

# Example usage within Streamlit
if 'acknowledged_warnings' not in st.session_state:
    st.session_state['acknowledged_warnings'] = False

if not st.session_state['acknowledged_warnings']:
    # Call your function here, assuming you have the 'workbook' and 'stream' variables set up
    # format_non_pivot_table(workbook, stream)
    pass  # Replace this with your actual function call
else:
    # Proceed with the rest of your Streamlit app
    st.write("Warnings acknowledged. Continuing with the app...")




#====================================================================================================================================================
# Function below will transform the Pivot table Spreadsheet into virtical columns and add the columns needed to import into Snowflake
#====================================================================================================================================================


def format_pivot_table(workbook, selected_option):
      # Assuming the sheet name is 'Sheet1', you can modify it as per your actual sheet name
    sheet = workbook['Sheet1']

    # Read the data from the sheet into a DataFrame
    data = sheet.values
    columns = next(data)  # Get the column names from the first row
    df = pd.DataFrame(data, columns=columns)

    # Get the store IDs from the column names
    store_ids = [x for x in df.columns[5:]]

    # Melt the data so that store IDs become a separate column
    df_melted = pd.melt(
        df,
        id_vars=df.columns[:5],
        value_vars=store_ids,
        var_name="store_id",
        value_name="Yes/No",
    )

    #st.write(df_melted.columns)

    # Replace 1 with a green checkmark and NaN with a red X
    df_melted['Yes/No'] = df_melted['Yes/No'].apply(lambda x: 'Yes' if x == 1 else ('No' if pd.isna(x) else '*'))
    #df_melted['Yes/No'] = df_melted['Yes/No'].apply(lambda x: 'Yes' if str(x).strip() == '1' else ('No' if str(x).strip().lower() == 'no' else 'No'))
    # Replace 1 with 'Yes' and NaN with 'No'
    #df_melted['Yes/No'] = df_melted['Yes/No'].apply(lambda x: 'Yes' if x == '1' else ('No' if pd.isna(x) else 'No'))

    




    # Move store_id column to the second position and rename it as STORE_NUMBER
    df_melted.insert(1, "STORE_NUMBER", df_melted.pop("store_id"))

    # Rename column "STORE NUMBER" to 'STORE NAME'
    df_melted.rename(columns={"STORE NUMBER": "STORE_NAME"}, inplace=True)
 
    # Add a new column "STORE_NAME" with empty values
    df_melted.insert(0, "STORE_NAME", "")

    # Reorder the columns with "STORE_NAME" in position 0, "STORE_NUMBER" in position 1, and "UPC" in position 2
    df_melted = df_melted[["STORE_NAME", "STORE_NUMBER", "UPC"] + [col for col in df_melted.columns if col not in ["STORE_NAME", "STORE_NUMBER", "UPC"]]]

  

    
    # Rename the columns as per your requirements
    df_melted.rename(columns={
        "Name": "PRODUCT_NAME",
        "Yes/No": "YES_NO",
        "SKU #": "SKU"
    }, inplace=True)

    # Display the updated DataFrame
    #print(df_melted)
    # Reindex the DataFrame with the desired columns
  
    
   

    # Remove ' and , characters from all columns
    df_melted = df_melted.replace({'\'': '', ',': '', '\*': '', 'Yes': '1', 'No': '0'}, regex=True)
    
    # Convert UPC entries to string, remove hyphens, and attempt to convert back to numbers
    df_melted['UPC'] = df_melted['UPC'].astype(str).str.replace('-', '', regex=True)
    temp_numeric_upc = pd.to_numeric(df_melted['UPC'], errors='coerce')  # Temporary numeric conversion for validation

    # Identify rows where UPC conversion failed using the temporary conversion data
    invalid_upc_rows = df_melted[temp_numeric_upc.isna()]

    if not invalid_upc_rows.empty:
        # Display an error and log the problematic rows
        st.error("Some UPC values could not be converted to numeric and may contain invalid characters or are empty. Please correct these in the original sheet and try uploading again.")
        st.dataframe(invalid_upc_rows[['UPC']])
        # Optionally, provide indices or additional info to help users locate the problem in their file
        st.write("Problematic row indices:", invalid_upc_rows.index.tolist())
        st.stop()  # Use this to stop further execution of the script

    # Since no issues, update UPC with its numeric version
    df_melted['UPC'] = temp_numeric_upc
    
    # # Convert UPC column to string
    # df_melted['UPC'] = df_melted['UPC'].astype(str)
    
    # # Remove '-' character from the UPC column
    # df_melted['UPC'] = df_melted['UPC'].str.replace('-', '')
    

  
    # Fill STORE_NAME column with "FOOD MAXX" starting from row 2
    df_melted.loc[0:, "STORE_NAME"] = st.session_state.selected_option

    # Fill SKU column with 0 starting from row 2
    df_melted.loc[0:, "SKU"] = 0
    #st.write(df_melted)
    df_melted.loc[0:,"ACTIVATION_STATUS"] =""
    df_melted.loc[0:,"COUNTY"] =""
    # Fill CHAIN_NAME column with "FOOD MAXX" starting from row 2
    df_melted.loc[0:, "CHAIN_NAME"] = st.session_state.selected_option
    #st.write(df_melted)

    
   

    # Convert DataFrame back to workbook object
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    for row in dataframe_to_rows(df_melted, index=False, header=True):
        new_sheet.append(row)

    return new_workbook


#====================================================================================================================================================
# END Function below will transform the Pivot table Spreadsheet into virtical columns and add the columns needed to import into Snowflake
#====================================================================================================================================================






#======================================================================================================================================
# This function will format the reset schedule to prepare for final integration into the reset_schedule table in Snowflake
#======================================================================================================================================
 
def format_RESET_schedule(workbook):
    # Get the worksheet by name ('Sheet1')
    ws = workbook['RESET_SCHEDULE_TEMPLATE']
    
    

    # # Fill column A with "CHAIN NAME SELECTED" starting from row 2
    # for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    #     for cell in row:
    #         cell.value = "FOOD MAXX"

    # Format column J as date "mm/dd/yyyy"
    date_format = NamedStyle(name='date_format', number_format='mm/dd/yyyy')
    for cell in ws['J']:
        cell.style = date_format
        

    # Loop through all cells in Column B
    for cell in ws['B']:
        # Check if the cell has a value (not empty)
        if cell.value:
            pass  # You can leave it empty or add specific logic for non-empty cells
        else:
            empty_cell_found = True
            st.warning("There is an empty Cell in column B (STORE_NUMBER). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
            return None  # Stop processing and return None when an empty cell is found

    # Initialize a flag to check for an empty cell
    empty_cell_found = False

    # Loop through all cells in Column C
    for cell in ws['C']:
        # Check if the cell has a value (not empty)
        if cell.value:
            # Convert the value to uppercase and remove leading/trailing spaces
            cell.value = cell.value.strip().upper()
        else:
            empty_cell_found = True
            st.warning("There is an empty Cell in column C (STORE_NAME). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
            return None  # Stop processing and return None when an empty cell is found
        
  




    # Initialize a flag to check for an empty cell
    empty_cell_found = False

    # Initialize a flag to check for an invalid date
    invalid_date_found = False

    # Regular expression pattern for date format (either 'mm/dd/yyyy' or 'm/d/yyyy')
    date_pattern = r'(\d{1,2}/\d{1,2}/\d{4})'

    # Loop through all cells in Column J, starting from row 2
    for cell in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        # Check if the cell has a value (not empty)
        if cell[0].value:
            # Check if the cell value is already a datetime object
            if not isinstance(cell[0].value, datetime):
                cell_value = str(cell[0].value)
                if not re.match(date_pattern, cell_value):
                    invalid_date_found = True
                    print(f"Invalid date format: {cell_value}")  # Print the invalid date value
                    break
        else:
            empty_cell_found = True
            print("Empty cell found")

    if invalid_date_found:
        st.warning("Invalid date format in column J (RESET_DATE). Please ensure all dates are correctly formatted as 'mm/dd/yyyy' or 'm/d/yyyy' and reformat the data.")
        return None  # Stop processing and return None when an invalid date is found

    if empty_cell_found:
        st.warning("There is an empty Cell in column J (RESET_DATE). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
        return None  # Stop processing and return None when an empty cell is found








    # Initialize a flag to check for an empty cell
    empty_cell_found = False

    # Loop through all cells in Column K (time)
    for cell in ws['K']:
        # Check if the cell has a value (not empty)
        if cell.value:
            pass
        else:
            empty_cell_found = True
            st.warning("There is an empty Cell in column K (TIME). Please resolve this issue then reformat again to ensure it will import into Snowflake.")
            return None  # Stop processing and return None when an empty cell is found

    # Rename Columns as required to meet the objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='CHAIN_NAME')
    ws.cell(row=1, column=2, value='STORE_NUMBER')
    ws.cell(row=1, column=3, value='STORE_NAME')
    ws.cell(row=1, column=4, value='PHONE')
    ws.cell(row=1, column=5, value='CITY')
    ws.cell(row=1, column=6, value='ADDRESS')
    ws.cell(row=1, column=7, value='STATE')
    ws.cell(row=1, column=8, value='COUNTY')
    ws.cell(row=1, column=9, value='TEAM_LEAD')
    ws.cell(row=1, column=10, value='RESET_DATE')
    ws.cell(row=1, column=11, value='TIME')
    ws.cell(row=1, column=12, value='STATUS')
    ws.cell(row=1, column=13, value='NOTES')

    # Return the workbook if no empty cells are found
    return workbook


#======================================================================================================================================
#END of function will format the reset schedule to prepare for final integration into the reset_schedule table in Snowflake
#======================================================================================================================================








