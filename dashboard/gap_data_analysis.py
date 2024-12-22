import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
import plotly.graph_objects as go
import snowflake.connector
from db_utils.snowflake_utils import get_snowflake_toml, create_gap_report, get_snowflake_connection

# Sidebar - Input for gap report parameters
st.sidebar.title("Gap Analysis Parameters")

# Create a connection to Snowflake
conn_toml = get_snowflake_toml(st.session_state.get('toml_info'))
cursor = conn_toml.cursor()

# Print current database and schema for debugging
cursor.execute("SELECT current_database(), current_schema()")
database_info = cursor.fetchone()
# st.write(f"Connected to database: {database_info[0]}, schema: {database_info[1]}")
# st.write(toml_info)
# Retrieve options from the database
try:
        
    salesperson_options = pd.read_sql('SELECT DISTINCT "SALESPERSON" FROM "SALESPERSON"', conn_toml)['SALESPERSON'].tolist()
except Exception as e:
    st.error(f"Error querying Salesperson table: {e}")


store_options = pd.read_sql("SELECT DISTINCT CHAIN_NAME FROM CUSTOMERS", conn_toml)['CHAIN_NAME'].tolist()
supplier_options = pd.read_sql("SELECT DISTINCT SUPPLIER FROM SUPPLIER_COUNTY", conn_toml)['SUPPLIER'].tolist()

salesperson_options.sort()
store_options.sort()
supplier_options.sort()

salesperson_options.insert(0, "All")
store_options.insert(0, "All")
supplier_options.insert(0, "All")

with st.sidebar.form(key="Gap Report Report", clear_on_submit=True):
    salesperson = st.selectbox("Filter by Salesperson", salesperson_options)
    store = st.selectbox("Filter by Chain", store_options)
    supplier = st.selectbox("Filter by Supplier", supplier_options)
    submitted = st.form_submit_button("Generate Gap Report")

df = None

with st.sidebar:
    if submitted:
        with st.spinner('Generating report...'):
            temp_file_path = create_gap_report(conn_toml, salesperson=salesperson, store=store, supplier=supplier)
            with open(temp_file_path, 'rb') as f:
                bytes_data = f.read()
            today = datetime.today().strftime('%Y-%m-%d')
            file_name = f"Gap_Report_{today}.xlsx"

            downloadcontainer = st.container()
            with downloadcontainer:
                st.download_button(label="Download Gap Report", data=bytes_data, file_name=file_name, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                st.write("File will be downloaded to your local download folder")

            container = st.container()
            with container:
                st.spinner('Generating report...')


# Function to format the sales report
def format_sales_report(workbook):
    try:
        # Format the sales report following the specified logic
        for sheet_name in workbook.sheetnames:
            if sheet_name != 'SALES REPORT':
                workbook.remove(workbook[sheet_name])

        ws = workbook['SALES REPORT']
        ws.delete_rows(2)
        ws.delete_cols(8)

        for cell in ws['F']:
            if cell.value is not None:
                cell.value = str(cell.value).replace('-', '')

        ws.insert_cols(2)
        ws.cell(row=1, column=2, value='STORE NAME')

        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell_offset = ws.cell(row=cell.row, column=2)
                cell_value = str(cell.value)
                store_name = cell_value.split('#')[0].replace("'", "") if '#' in cell_value else cell_value.replace("'", "")
                cell_offset.value = store_name

        ws.delete_cols(3)

        for col in ['B', 'E', 'C']:
            for cell in ws[col]:
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.replace(',', ' ').replace(" 's", "").replace("'", "")

        for cell in ws['F']:
            if cell.value is not None:
                cell.value = str(cell.value).replace('Is Null', '0')

        for cell in ws['G'][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER
            elif isinstance(cell.value, str):
                cell.number_format = numbers.FORMAT_NUMBER
                try:
                    cell.value = float(cell.value.replace(",", ""))
                except ValueError:
                    pass

        return workbook

    except Exception as e:
        st.error(f"Error occurred while formatting the sales report: {str(e)}")
        return None

# Upload the workbook
uploaded_file = st.file_uploader(":red[Upload freshly ran sales table from your application]", type=["xlsx", "xls"])

if uploaded_file is not None:
    workbook = load_workbook(uploaded_file)
    if st.button("Reformat"):
        new_workbook = format_sales_report(workbook)
        if new_workbook:
            new_filename = 'formatted_' + uploaded_file.name
            stream = BytesIO()
            new_workbook.save(stream)
            stream.seek(0)
            st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
        else:
            st.error("Failed to reformat the uploaded file.")

# Function to write sales report data to Snowflake
def write_salesreport_to_snowflake(df):
    try:
        df.fillna(value="NULL", inplace=True)

        toml_info = st.session_state.get('toml_info')
        if not toml_info:
            st.error("TOML information is not available. Please check the tenant ID and try again.")
            return

        # Get Snowflake connection
        conn_toml = get_snowflake_toml(toml_info)
        if not conn_toml:
            return

        # Create a cursor object
        cursor = conn_toml.cursor()

        current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        sql_query = f"""
        CREATE OR REPLACE TABLE SALES_REPORT AS
        SELECT 
            CAST(STORE_NUMBER AS NUMBER) AS STORE_NUMBER, 
            CAST(TRIM(STORE_NAME) AS VARCHAR) AS STORE_NAME, 
            CAST(ADDRESS AS VARCHAR) AS ADDRESS, 
            CAST(SALESPERSON AS VARCHAR) AS SALESPERSON, 
            CAST(PRODUCT_NAME AS VARCHAR) AS PRODUCT_NAME, 
            CAST(UPC AS NUMERIC) AS UPC, 
            CAST(PURCHASED_YES_NO AS NUMERIC) AS PURCHASED_YES_NO,
            CAST('{current_timestamp}' AS TIMESTAMP) AS LAST_UPLOAD_DATE
        FROM (VALUES {', '.join([str(tuple(row)) for row in df.values])}) 
        AS tmp(STORE_NUMBER, STORE_NAME, ADDRESS, SALESPERSON, PRODUCT_NAME, UPC, PURCHASED_YES_NO);
        """

        cursor.execute(sql_query)
        conn_toml.commit()

        st.success("Data has been imported into Snowflake table SALES_REPORT!")

    except snowflake.connector.errors.Error as e:
        st.error(f"Error writing to Snowflake: {str(e)}")
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn_toml' in locals():
            conn_toml.close()

# Upload formatted sales report to write to Snowflake
uploaded_file = st.file_uploader(":red[UPLOAD CURRENT SALES REPORT AFTER IT HAS BEEN FORMATTED]", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading Sales Report Data to Snowflake ...'):
            write_salesreport_to_snowflake(df)

# Function to create and display the gap analysis bar chart
def gap_analysis_bar_chart():
    toml_info = st.session_state.get('toml_info')
    if not toml_info:
        st.error("TOML information is not available. Please check the tenant ID and try again.")
        return

    # Create a connection to Snowflake
    conn_toml = get_snowflake_toml(toml_info)

    # Create a cursor object
    cursor = conn_toml.cursor()

    # Retrieve data from your view
    query = """
        SELECT 
            SUM("In_Schematic") AS total_in_schematic, 
            SUM("PURCHASED_YES_NO") AS purchased, 
            SUM("PURCHASED_YES_NO") / COUNT(*) AS purchased_percentage 
        FROM GAP_REPORT;
    """
    df = pd.read_sql(query, conn_toml)

    # Check if the DataFrame is not empty and has the expected columns
    expected_columns = ['TOTAL_IN_SCHEMATIC', 'PURCHASED', 'PURCHASED_PERCENTAGE']
    if df.empty:
        st.error("No data retrieved from GAP_REPORT.")
        return

    if not all(col in df.columns for col in expected_columns):
        st.error(f"Unexpected columns in the retrieved data. Expected columns: {expected_columns}, but got: {df.columns.tolist()}")
        return

    # Format the 'PURCHASED_PERCENTAGE' column as a percentage with 2 decimal places
    df['PURCHASED_PERCENTAGE'] = (df['PURCHASED'] / df['TOTAL_IN_SCHEMATIC'] * 100).map('{:.2f}%'.format)

    # Create the bar chart
    fig = go.Figure(data=[
        go.Bar(
            x=['Total in Schematic', 'Purchased', 'Purchased Percentage'], 
            y=[df['TOTAL_IN_SCHEMATIC'].iloc[0], df['PURCHASED'].iloc[0], df['PURCHASED_PERCENTAGE'].iloc[0]], 
            text=[df['TOTAL_IN_SCHEMATIC'].iloc[0], df['PURCHASED'].iloc[0], df['PURCHASED_PERCENTAGE'].iloc[0]], 
            textposition='auto', 
            marker=dict(color=['#1f77b4', '#ff7f0e', '#2ca02c'])
        )
    ])

    # Set the axis labels and plot title
    fig.update_layout(
        xaxis_title='',
        yaxis_title='Number of Items',
        title='Total Items in Schematic vs. Purchased Items',
        plot_bgcolor='#B3D7ED',  # Set the background color
        paper_bgcolor='#F8F2EB' # Set the paper (border) color
    )
  
    # Add a border to the chart
    fig.update_traces(
        marker_line_width=1.5,
        marker_line_color='black'
    )

    # Customize the bar chart colors
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c']
    for i in range(len(fig.data)):
        fig.data[i].marker.color = colors[i]
        fig.data[i].marker.line.width = 1.5
        fig.data[i].marker.line.color = 'black'

    # Row A
    col1, col2 = st.columns(2)

    with col1:
        container = st.container()
        with container:
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        container = st.container()
        with container:
            st.plotly_chart(fig, use_container_width=True)
conn_toml.close()
    # except Exception as e:
    #     st.error(f"Failed to generate gap analysis chart: {str(e)}")
    # finally:
    #     if 'conn_toml' in locals():
    #         conn_toml.close()

gap_analysis_bar_chart()
